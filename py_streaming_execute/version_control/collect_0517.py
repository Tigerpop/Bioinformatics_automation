# -*- coding: utf-8 -*-
import re,os,sys,shutil,chardet
import pandas as pd
from pandas import DataFrame
from datetime import date
import warnings
warnings.filterwarnings('ignore')

sample = sys.argv[1] # 第一个参数
sample_path = sys.argv[2]
log_path = sys.argv[3]
generate_location = sys.argv[4]
bed_key = sys.argv[5]

def generate_summary():
    # 实现多个sheet 写入一个exel。
    with pd.ExcelWriter(f'{generate_location}/{sample_path}/{sample}.summary.xlsx') as writer:
        try:
            # mate
            with open('/received/main/received.csv', 'rb') as f0: # 确认编码类型。
                encoding_stype = chardet.detect(f0.read())
            df_receive = pd.read_csv(f'/received/main/received.csv',sep=',',header=1,encoding=encoding_stype['encoding'])
            df_mate = df_receive[df_receive['样本编号*']==sample]
            df_mate[['id','name','gender','age','cancer','clinname','送检医院','panel','projectname','报告模版','样本类型*','到样日期*']] \
            = df_mate[['样本编号*','姓名*','性别','年龄','肿瘤类型*','临床诊断*','送检医院','探针*','检测项目*','报告模板*','样本类型*','到样日期*']]
            df_mate = df_mate[['id','name','gender','age','cancer','clinname','送检医院','到样日期*','样本类型*','panel','projectname','报告模版']]
            # df_mate['id'] = df_mate['id'].apply(lambda x: x[:-2])
            df_mate = df_mate.tail(1)
            DataFrame(df_mate).to_excel(writer,sheet_name='meta',index=False,header=True)
            
            # snpindel 
            file = f'{generate_location}/{sample_path}/{sample}.process.hg19_multiprocess.txt'
            df_snpindel = pd.read_csv(file,sep='\t')
            df_snpindel['review'] = None
            DataFrame(df_snpindel).to_excel(writer,sheet_name='snpindel',index=False,header=True)

            # snpindel_germline 
            file = f'{generate_location}/{sample_path}/{sample}.germline.hg19_multigermline.txt'
            df_germline = pd.read_csv(file,sep='\t')
            df_germline['review'] = None
            DataFrame(df_germline).to_excel(writer,sheet_name='germline',index=False,header=True)
            
            
            # fusion
            file = f"{generate_location}/{sample_path}/factera_generate/{sample}.bwa_mem.factera.fusions.txt"
            head = ['Est_Type','Region1','Region2','Break1','Break2','Break_support1','Break_support2','Break_offset','Orientation','Order1','Order2','Break_depth','Proper_pair_support','Unmapped_support','Improper_pair_support','Paired_end_depth','Total_depth','Fusion_seq','<>','Non-templated_seq','-']
            df_fusion = pd.read_csv(file,sep='\t',header=None,names=head,skiprows=1)
            df_fusion['review'] = None
            DataFrame(df_fusion).to_excel(writer,sheet_name='fusion',index=False)
            
            # cnv
            # file = f'{generate_location}/{sample_path}/decon_generate/DECoNtestCalls_all.txt'
            # if not os.path.exists(file):
            #     head = ['CNV.ID','Sample','Correlation','N.comp','Start.b','End.b','CNV.type','N.exons','Start','End','Chromosome','Genomic.ID','BF','Reads.expected','Reads.observed','Reads.ratio','Gene','Custom.first','Custom.last','cnv.number','review']
            #     df_cnv = pd.DataFrame(columns=head)
            #     DataFrame(df_cnv).to_excel(writer,sheet_name='cnv',index=False,header=True)
            # else:
            #     df_cnv = pd.read_csv(file,sep='\t')
            #     df_cnv['cnv.number'] = df_cnv['Reads.ratio']*2 
            #     df_cnv['review'] = None
            #     DataFrame(df_cnv).to_excel(writer,sheet_name='cnv',index=False,header=True)
            file = f'{generate_location}/{sample_path}/panelcn_generate/panelcn_output.txt'
            if not os.path.exists(file):
                head = ['Sample','Chr','Gene','Exon','Start','End','RC','medRC','RC.norm','medRC.norm','lowQual','CN','review']
                df_cnv = pd.DataFrame(columns=head)
                DataFrame(df_cnv).to_excel(writer,sheet_name='cnv',index=False,header=True)
            else:
                df_cnv = pd.read_csv(file,sep='\t')
                df_cnv['CNN'] = (df_cnv['RC.norm'].astype(int)*2/df_cnv['medRC.norm'].astype(int)).astype(int)
                df_cnv['review'] = None
                DataFrame(df_cnv).to_excel(writer,sheet_name='cnv',index=False,header=True)
            
            # msi 和 chemo
            if bed_key in ['Q120','SD160','NBC650','BCP650']: 
                file1 = f"{generate_location}/{sample_path}/msi_generate/msi_result"
                file2 = f"{generate_location}/{sample_path}/msi_generate/msi_result_unstable"
                file = f"{generate_location}/{sample_path}/msi_generate/to_summery_smi"
                with open(file1,'r')as f0,open(file2,'r')as f1,open(file,'w')as f2:
                    for line in f0:
                        f2.write(line)
                    for line in f1:
                        f2.write(line)
                df_msi = pd.read_csv(file,sep='\t',header=None,names=[1,2,3,4,5,6,7,8,9,10])
                DataFrame(df_msi).to_excel(writer,sheet_name='msi',index=False,header=False)
                file = f"{generate_location}/{sample_path}/chemo_generate/process_output_base_num.txt"
                df_chemo = pd.read_csv(file,sep='\t')
                df_chemo['review'] = None
                DataFrame(df_chemo).to_excel(writer,sheet_name='chemo',index=False,header=True)
            '''
            # hla 和 neoantigen
            if bed_key in ['NBC650','BCP650']: 
                file = f'{generate_location}/{sample_path}/optitype_generate/fished_result.tsv'
                df_hla = pd.read_csv(file,sep='\t')
                df_hla = df_hla.rename(columns={'Unnamed: 0':'Patient'})      # 重命名一下第一列。
                df_hla['Patient'] = sample
                DataFrame(df_hla).to_excel(writer,sheet_name='hla',index=False,header=True) 
                file = f'{generate_location}/{sample_path}/neoantigen_generate/{sample}.format_neoantigens.txt'
                df_neoantigen = pd.read_csv(file,sep='\t')
                df_neoantigen['review'] = None
                DataFrame(df_neoantigen).to_excel(writer,sheet_name='neoantigen',index=False,header=True) 
            '''
            # qc   
            file = f"{log_path}/quality_control/Quality_Control.txt"
            file1 = f"{log_path}/temp_qc"
            with open(file, 'r') as f:
                lines = f.readlines()
            non_empty_lines = [line.strip() for line in lines if line.strip()]
            dict = {}
            for line in non_empty_lines:
                if len(line.split(":"))>1:
                    dict[line.split(":")[0]] = line.split(":")[1]
                # else:
                #      dict[line.split(":")[0]] = ' '
            with open(file1,'w')as f:
                print(dict.values())
                f.write('\t'.join(dict.keys())+'\n')
                f.write('\t'.join(dict.values())+'\n')
            df_qc = pd.read_csv(file1,sep='\t')
            DataFrame(df_qc).to_excel(writer,sheet_name='qc',index=False,header=True)
            
        except Exception as e:
            # print(e, type(e))
            if (isinstance(e, pd.errors.EmptyDataError)):
                print(f"这里对空行文件\n{file}\n进行处理")
                if re.search('fusion',file)!=None:
                    head = ['Est_Type','Region1','Region2','Break1','Break2','Break_support1','Break_support2','Break_offset','Orientation','Order1','Order2','Break_depth','Proper_pair_support','Unmapped_support','Improper_pair_support','Paired_end_depth','Total_depth','Fusion_seq','<>','Non-templated_seq','-','review']
                    df_fusion = pd.DataFrame(columns=head)
                    DataFrame(df_fusion).to_excel(writer,sheet_name='fusion',index=False,header=True)
            sys.stderr.write(f'出现错误：\n {e} \n')
            exit(1) # 让上一层看见 这一步出现的错误。
            
def send_summary_to_archive(): # 和之前批处理不一样，这里只需要把这一个sample 添加进archive文件夹的当天子文件夹内。
    try:
        today=date.today().strftime("%Y%m%d")
        summary_file = f'{generate_location}/{sample_path}/{sample}.summary.xlsx'
        print(summary_file)
        today=date.today().strftime("%Y%m%d")
        if not os.path.exists(f'/archive/{today}'):
            os.mkdir(f'/archive/{today}')
        shutil.copyfile(summary_file, f'/archive/{today}/{sample}.summary.xlsx')
    except Exception as e:
        sys.stderr.write(f'出现错误：\n {e} \n')
        exit(1) # 让上一层看见 这一步出现的错误。
        
# def add_review_lable(input_path=f'{generate_location}/{sample_path}/{sample}.summary.xlsx',sheet1='snpindel',sheet2='germline'):
#     df = pd.read_excel(input_path, sheet_name=None)
#     df0 = df[sheet1]
#     
#     df1 = df['temp_sheet']
#     df1['vaf_temp'] = df1['VAF'].str.replace('%', '').astype(float)
#     df1 = df1.sort_values('vaf_temp', ascending=False)
#     df1 = df1.drop(columns=['vaf_temp'])
#     df_appended = df0.append(df1, ignore_index=True)
#     workbook.remove(worksheet)
#     workbook.remove(temp_sheet)
#     workbook.save(output)# ('add_color.xlsx')

    # with pd.ExcelWriter(output,mode='a') as writer:
    #     df_appended.to_excel(writer,sheet_name=sheet,index=False,header=True)
        

##########################################################################################
# 以下为加颜色部分；
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors
import re
import pandas as pd
from pandas import DataFrame

def add_color(input,output,sheet='snpindel'):
    # 打开 Excel 文件
    workbook = load_workbook(input) # ('./BCX-YCH0784-17T-1G0507.summary.xlsx')
    worksheet = workbook[sheet]

    red_rows = []
    orange_rows = []
    green_rows = []
    brown_rows = []
    none_rows = []
    header = []
    for cell_ in next(worksheet.iter_rows()):
        header.append(cell_.value)

    # 遍历每一行
    # min_row 和 min_col 是 iter_rows 方法的参数，用于指定要迭代的单元格范围的起始行和起始列。
    # 默认情况下，min_row 和 min_col 的值为 1，表示从第一行和第一列开始迭代。
    for row in worksheet.iter_rows(min_row=2, min_col=1):
        for i,cell_ in enumerate(row):
            #print(header[i])
            if header[i] == 'CLNSIG':
                cell_value_str = str(cell_.value)
                result = re.findall('Pathogenic|pathogenic|Likely_pathogenic',cell_value_str)
                if result:
                    for cell in row:
                        cell.font = Font(color='DF0F18' )
                    result = re.findall('Conflict.*pathogenicity|conflict.*pathogenicity', cell_value_str)
                    if result:
                        for cell in row:
                            cell.font = Font(color='EB7621')
                        orange_rows.append(row)
                    else:
                        red_rows.append(row)
                    break
                result = re.findall('Conflict.*pathogenicity|conflict.*pathogenicity|Uncertain|uncertain|not_provided',cell_value_str) # drug_response
                if result:
                    for cell in row:
                        cell.font = Font(color='EB7621' )
                    orange_rows.append(row)
                    break
                result = re.findall('Benign|benign',cell_value_str)
                if result:
                    for cell in row:
                        cell.font = Font(color='0F7F12' )
                    green_rows.append(row)
                    break
                result = re.findall('drug_response|Drug_response',cell_value_str)
                if result:
                    for cell in row:
                        cell.font = Font(color='B77346' )
                    brown_rows.append(row)
                    break
        # 如果 CLNSIG 不包含以上几个关键词，则将该行加入到 none_rows 列表中
        else:
            none_rows.append(row)


    # CLNSIG 有值提前
    first_row = 2
    last_row = worksheet.max_row
    worksheet.delete_rows(2, worksheet.max_row)

    for row in red_rows:
        worksheet.append(row)

    for row in orange_rows:
        worksheet.append(row)
        
    for row in brown_rows:
        worksheet.append(row)

    for row in green_rows:
        worksheet.append(row)



    # for row in none_rows:
    #     worksheet.append(row)

    # 保存文件
    temp_sheet = workbook.create_sheet('temp_sheet')
    temp_sheet.append([cell.value for cell in workbook[sheet][1]]) # 第一行写入
    for row in none_rows:
        temp_sheet.append([cell.value for cell in row])
    workbook.save(f'{generate_location}/{sample_path}/add_color.xlsx')

    df = pd.read_excel(f'{generate_location}/{sample_path}/add_color.xlsx', sheet_name=None)
    df0 = df[sheet]
    df1 = df['temp_sheet']
    df1['vaf_temp'] = df1['VAF'].str.replace('%', '').astype(float)
    df1 = df1.sort_values('vaf_temp', ascending=False)
    df1 = df1.drop(columns=['vaf_temp'])
    df_appended = df0.append(df1, ignore_index=True)
    workbook.remove(worksheet)
    workbook.remove(temp_sheet)
    workbook.save(output)# ('add_color.xlsx')

    with pd.ExcelWriter(output,mode='a') as writer:
        df_appended.to_excel(writer,sheet_name=sheet,index=False,header=True)

def add_color1(input,output,sheet='snpindel'):
    # 打开 Excel 文件
    workbook = load_workbook(input) # ('./BCX-YCH0784-17T-1G0507.summary.xlsx')
    worksheet = workbook[sheet]

    red_rows = []
    orange_rows = []
    green_rows = []
    brown_rows = []
    none_rows = []
    header = []
    for cell_ in next(worksheet.iter_rows()):
        header.append(cell_.value)

    # 遍历每一行
    # min_row 和 min_col 是 iter_rows 方法的参数，用于指定要迭代的单元格范围的起始行和起始列。
    # 默认情况下，min_row 和 min_col 的值为 1，表示从第一行和第一列开始迭代。
    for row in worksheet.iter_rows(min_row=2, min_col=1):
        for i,cell_ in enumerate(row):
            #print(header[i])
            if header[i] == 'CLNSIG':
                cell_value_str = str(cell_.value)
                result = re.findall('Pathogenic|pathogenic|Likely_pathogenic',cell_value_str)
                if result:
                    for cell in row:
                        cell.font = Font(color='DF0F18' )
                    result = re.findall('Conflict.*pathogenicity|conflict.*pathogenicity', cell_value_str)
                    if result:
                        for cell in row:
                            cell.font = Font(color='EB7621')
                        orange_rows.append(row)
                    else:
                        red_rows.append(row)
                    break
                result = re.findall('Conflict.*pathogenicity|conflict.*pathogenicity|Uncertain|uncertain|not_provided',cell_value_str) # drug_response
                if result:
                    for cell in row:
                        cell.font = Font(color='EB7621' )
                    orange_rows.append(row)
                    break
                result = re.findall('Benign|benign',cell_value_str)
                if result:
                    for cell in row:
                        cell.font = Font(color='0F7F12' )
                    green_rows.append(row)
                    break
                result = re.findall('drug_response|Drug_response',cell_value_str)
                if result:
                    for cell in row:
                        cell.font = Font(color='B77346' )
                    brown_rows.append(row)
                    break
        # 如果 CLNSIG 不包含以上几个关键词，则将该行加入到 none_rows 列表中
        else:
            none_rows.append(row)


    # CLNSIG 有值提前
    first_row = 2
    last_row = worksheet.max_row
    worksheet.delete_rows(2, worksheet.max_row)

    for row in red_rows:
        worksheet.append(row)

    for row in orange_rows:
        worksheet.append(row)

    for row in brown_rows:
        worksheet.append(row)

    for row in green_rows:
        worksheet.append(row)

    for row in none_rows:
        worksheet.append(row)

    # 保存文件
    workbook.save(output)# ('add_color.xlsx')

def order_and_rename(input):
    sheet_names = ['meta', 'snpindel', 'germline','fusion','cnv','qc']  # 按您想要的顺序排列工作表
    workbook = load_workbook(input)
    sheets = [workbook[sheet_name] for sheet_name in sheet_names]
    for idx, sheet in enumerate(sheets):
        workbook.move_sheet(sheet, idx + 1)
    workbook.save(input)

def add_color0(input='./BCX-YCH0784-17T-1G0507.summary.xlsx',sheet='cnv'):
    df = pd.read_excel(input, sheet_name=None)
    df = df[sheet]
    # print(df['CN'])
    # 筛选出CN列等于"CN4"的行中 gene 出现的行。
    cn4_rows = df[df['CN'] == 'CN4']
    cn4_rows_gene = cn4_rows['Gene'].tolist()
    cn4_rows_gene_list = list(set(cn4_rows_gene))
    cn4_rows = df[df['Gene'].isin(cn4_rows_gene_list)]
    cn4_rows = cn4_rows.sort_values(by=['Gene'])
    # print(cn4_rows)
    # print(cn4_rows_gene_list)
    # 将这些行移到表的最上方
    df = pd.concat([cn4_rows, df.drop(cn4_rows.index)])

    workbook = load_workbook(input) # ('./BCX-YCH0784-17T-1G0507.summary.xlsx')
    worksheet = workbook[sheet]
    workbook.remove(worksheet)
    workbook.save(input)# ('add_color.xlsx')

    with pd.ExcelWriter(input,mode='a') as writer:
        df.to_excel(writer,sheet_name=sheet,index=False,header=True)

    # 设置字体为红色
    workbook = load_workbook(input) # ('./BCX-YCH0784-17T-1G0507.summary.xlsx')
    worksheet = workbook[sheet]
    header = []
    for cell_ in next(worksheet.iter_rows()):
        header.append(cell_.value)
    for row in worksheet.iter_rows(min_row=2, min_col=1):
        for i,cell_ in enumerate(row):
            if header[i] == 'Gene':
                if cell_.value in cn4_rows_gene_list:
                    for cell in row:
                        cell.font = Font(color='DF0F18' )
    # 保存文件
    workbook.save(input)# ('add_color.xlsx')

def add_color_rename_resort(input_path=f'{generate_location}/{sample_path}/{sample}.summary.xlsx'):
    add_color0(input=input_path)
    add_color(input=input_path, output=input_path,sheet='snpindel')
    add_color1(input=input_path, output=input_path,sheet='snpindel')
    add_color(input=input_path, output=input_path,sheet='germline')
    add_color1(input=input_path, output=input_path,sheet='germline')
    order_and_rename(input=input_path)

if __name__=='__main__':
    generate_summary()
    # add_review_lable()
    add_color_rename_resort()
    send_summary_to_archive()
